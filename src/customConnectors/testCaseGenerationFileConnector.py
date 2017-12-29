'''
Created on Nov 9, 2017

@author: IBM
'''


import os
import xlsxwriter
import pandas as pd


# from time import sleep
# from datetime import datetime

from openpyxl import load_workbook
from baseUtils import testCaseGenerationLogger
from baseUtils import testCaseGenerationErrorHandler   
from mainController import testCaseGenerationConfig




class testCaseGenerationFileConnector(object):

   
        
    
    def __init__(self):
      logger =   testCaseGenerationLogger.testCaseGenerationLogger()
      self.log = logger.getLogger(self.__class__.__name__)
      #self.log.info('This is logger for testCaseGenerationFileConnector:init method')
      
      
   


    def getxls_Connection(self,userName):
      fileError=testCaseGenerationErrorHandler.testCaseGenerationErrorHandler()   
        
      try:  
        filename=os.path.join(testCaseGenerationConfig.testcase_destination_path,testCaseGenerationConfig.testcase_destination_filemask)+ "_" + userName + ".xlsx"
        print filename    
        if os.path.exists(filename):
           os.remove(filename)
           print "File removed.."
        else:
           print ("File is not there")
           
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet('TestCases')
#         worksheet.write(0,0,"Mapping Reference Number")
#         worksheet.write(0,1,"Action")
#         worksheet.write(0,2,"ProcessCategory")
#         worksheet.write(0,3,"Query Type")
#         worksheet.write(0,4,"Test case Category")
#         worksheet.write(0,5,"Expected Result")
#         worksheet.write(0,6,"Query Execution Area")
#         worksheet.write(0,7,"Actual Result")
#         worksheet.write(0,8,"Execution Date")
#         worksheet.write(0,9,"Comments")

        worksheet.write(0,0,"Test Case No")
        worksheet.write(0,1,"Mapping Reference Number")
        worksheet.write(0,2,"Test case Category")
        worksheet.write(0,3,"Test Case Description")
        worksheet.write(0,4,"Query Execution Area")
        worksheet.write(0,5,"Action")
        worksheet.write(0,6,"Expected Result")
#         worksheet.write(0,7,"ProcessCategory")
#         worksheet.write(0,8,"Query Type")
        worksheet.write(0,7,"Actual Result")
        worksheet.write(0,8,"Execution Date")
        worksheet.write(0,9,"Comments")
        
    

        
        worksheet1 = workbook.add_worksheet('Classification')
        worksheet1.write(0,0,"ProcessCategory")
        worksheet1.write(0,1,"CountOfMapping")
        worksheet1.write(0,2,"TestCaseCount")
#         worksheet1.write(0,3,"Tracebility IDs")
#         
        workbook.close()

        print ("start Creating File")
        book = load_workbook(filename)
        xls_tuple = (filename,book)
        return xls_tuple
        #return book
      except Exception as e:
           exceptionmsg=fileError.check_Technical_Exception(e)
           print exceptionmsg
           self.log.error(exceptionmsg)
           exit()
           


