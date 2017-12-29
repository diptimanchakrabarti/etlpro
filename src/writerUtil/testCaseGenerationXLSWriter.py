'''
Created on Nov 9, 2017

@author: IBM
'''

from pandas import ExcelFile

from pandas import ExcelWriter
import pandas as pd
import os

from baseUtils.testCaseGenerationLogger import testCaseGenerationLogger
from customConnectors import testCaseGenerationFileConnector
from baseUtils import testCaseGenerationErrorHandler
from mainController import testCaseGenerationConfig
import openpyxl
from openpyxl.styles import Font,PatternFill
from openpyxl.writer.excel import save_virtual_workbook

class testCaseGenerationXLSWriter(object):

   
 
    def __init__(self):
#       filename="TestScript.xlsx"   
      logger =   testCaseGenerationLogger()
      self.log = logger.getLogger(self.__class__.__name__)
      #self.log.info('This is logger for testCaseGenerationFileConnector:init method')
      self.errhandler=testCaseGenerationErrorHandler.testCaseGenerationErrorHandler()
      self.fc = testCaseGenerationFileConnector.testCaseGenerationFileConnector()
      self.xls_tuple = self.fc.getxls_Connection(testCaseGenerationConfig.user_name)

    def write_xls(self,df):
        final_comparedf_default=pd.DataFrame([])
        final_basedf_default=pd.DataFrame([])
        
        #book = fc.getxls_Connection('TestScript.xlsx')
        fileName = self.xls_tuple[0]
#         print fileName 
        book = self.xls_tuple[1]
#       print book
        writer=pd.ExcelWriter(fileName, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#         for ws in book.worksheets:
#           print ws
        #df1.to_excel(writer,'SOURCE', startrow=1, header=None, index=False)
        #df2 = df1
        #df2.to_excel(writer,'TARGET', startrow=1, header=None, index=False)
#       print df["TestCase"]
        try:
#             comparedf_default = df[(df['ProcessCategory']=='DEFAULT')  & (df['TargetXLSTab']=='SOURCE')]
#             final_comparedf_default=comparedf_default.filter(items=['Tracebility ID','TestCase','ProcessCategory'])
#             basedf_default= df[(df['ProcessCategory']=='DEFAULT')  & (df['TargetXLSTab']=='TARGET')]
#             final_basedf_default=basedf_default.filter(items=['Tracebility ID','TestCase','ProcessCategory'])
#             
#             final_comparedf_default.to_excel(writer,'SOURCE', startrow=1, header=None, index=False)
#             final_basedf_default.to_excel(writer,'TARGET', startrow=1, header=None, index=False)
            
#             df_source=df[(df['TargetXLSTab']=='SOURCE')] 
#             df_target=df[(df['TargetXLSTab']=='TARGET')]
#             
#             df_source=df_source.sort_values(by=['ProcessCategory','Tracebility ID'])
#             df_target=df_target.sort_values(by=['ProcessCategory','Tracebility ID'])
#             
#             df_source=df_source.filter(items=['Tracebility ID','TestCase','ProcessCategory','Query Type'])
#             df_target=df_target.filter(items=['Tracebility ID','TestCase','ProcessCategory','Query Type'])
#             
#             df_source.to_excel(writer,'SOURCE', startrow=1, header=None, index=False)
#             df_target.to_excel(writer,'TARGET', startrow=1, header=None, index=False)


#            df=df.sort_values(by=['ProcessCategory','Tracebility ID'])
            #df=df.filter(items=['Mapping Reference Number','Action','ProcessCategory','Query Type','Test case Category','Expected Result','Query Execution Area'])
            df=df.filter(items=['Test Case No','Mapping Reference Number','Test case Category','Test Case Description','Query Execution Area','Action','Expected Result'])
            
            df=df.sort_values('Test Case No')
            df.to_excel(writer,'TestCases', startrow=1, header=None, index=False)
            writer.save()
            final_filename=self.format_excel(fileName)
#             self.format_excel(fileName)
            return final_filename
            self.log.info('testcase file created successfully')

        except Exception as e:
            exceptionmsg=self.errhandler.check_Technical_Exception(e)
            self.log.error(exceptionmsg)
            print exceptionmsg
    
    def write_xls_classification(self,df):

#         fc = testCaseGenerationFileConnector.testCaseGenerationFileConnector()
#         xls_tuple = fc.getxls_Connection(userName)
        #book = fc.getxls_Connection('TestScript.xlsx')
        fileName = self.xls_tuple[0]
#       print fileName 
        book = self.xls_tuple[1]
#       print book
        writer=pd.ExcelWriter(fileName, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        try:
            df.to_excel(writer,'Classification', startrow=1, header=None, index=False)
            writer.save() 
            self.log.info('Testcase classification generated.')
        except Exception as e:
            exceptionmsg=self.errhandler.check_Technical_Exception(e)
            self.log.error(exceptionmsg)
            print exceptionmsg               
    
    def format_excel(self,fname):
        wb = openpyxl.load_workbook(fname)
        ws = wb.active
        #print ws
        row_num_base=2
        
#         print ws.max_row
        while (row_num_base < ws.max_row):
            start_pos ='A'+str(row_num_base)
            pass
            start_posB='B'+str(row_num_base)
            start_posC='C'+str(row_num_base)
            start_posD='D'+str(row_num_base)
            start_posG='G'+str(row_num_base)
#             match_end=0
            for row_num in range((row_num_base+1),ws.max_row+1):
                next_cell='A'+str(row_num)
                if (ws[start_pos].value == ws[next_cell].value):
                    end_pos='A'+str(row_num)
                    end_posB='B'+str(row_num)
                    end_posC='C'+str(row_num)
                    end_posD='D'+str(row_num)
                    end_posG='G'+str(row_num)
                else:
#                     match_end=1
                    break
            merge_set=start_pos+":"+   end_pos
            merge_setC=start_posC+":"+end_posC
            merge_setD=start_posD+":"+end_posD
            merge_setG=start_posG+":"+end_posG
            merge_setB=start_posB+":"+end_posB
#             print merge_set
            ws.merge_cells(merge_set)
            ws.merge_cells(merge_setB)
            ws.merge_cells(merge_setC)
            ws.merge_cells(merge_setD)
            ws.merge_cells(merge_setG)
                #ws.merge_cells('A2:A3')
            row_num_base=row_num
            #print row_num_base
#         wb.save(fname)
#         temp_fname=os.path.join(testCaseGenerationConfig.testcase_destination_path,'temp_file.xlsx')
        final_fname=os.path.join(testCaseGenerationConfig.testcase_destination_path,testCaseGenerationConfig.testcase_destination_filemask)+ "_" + "Final" + ".xlsx"
        fontObj = Font(name='Calibri', bold=True,color='FFFFFFFF')
        redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
        for col in range(1,ws.max_column+1):
            ws.cell(row=1,column=col).font= fontObj
            ws.cell(row=1,column=col).fill = redFill
        wb.save(final_fname)
        return final_fname
#         return save_virtual_workbook(wb)
        #os.rename(final_fname, 'C:\\Users\\IBM_ADMIN\\workspacepython\\testCaseGeneratorUtilProject\\src\\mainController\\templates\\TestScript_Final.xlsx')
#         return final_fname
#         os.remove(fname)
#         os.rename(temp_fname,fname)
